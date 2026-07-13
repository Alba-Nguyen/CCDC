from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.properties import CalcProperties


ROOT = Path.cwd()
STAGE = ROOT / "TOOL_Check_CCDC" / "_temp_stage.xlsx"
wb = openpyxl.load_workbook(STAGE, data_only=False)
bkct = wb["02.BKCT"]
bkpb = wb["03.BKPB"]


def last_detail(ws, start: int, key_cols: tuple[int, ...]) -> int:
    for row in range(ws.max_row, start - 1, -1):
        if any(ws.cell(row, col).value not in (None, "") for col in key_cols):
            return row
    return start


bkct_last = last_detail(bkct, 9, (3, 5, 6))
bkpb_last = last_detail(bkpb, 6, (1, 2, 6, 10))

for sheet_name, account in (("CheckChiTiet_2421", 2421), ("CheckChiTiet_2422", 2422)):
    ws = wb[sheet_name]
    last = last_detail(ws, 4, (1, 2))
    ws["D1"] = account
    ws["D1"].font = Font(name="Arial", size=16, bold=True, color="000000")
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center")

    for r in range(4, last + 1):
        ws.cell(r, 3).value = f'=IF($A{r}="",SUMIFS(\'02.BKCT\'!$F$9:$F${bkct_last},\'02.BKCT\'!$AD$9:$AD${bkct_last},$B{r},\'02.BKCT\'!$E$9:$E${bkct_last},$D$1),SUMIFS(\'02.BKCT\'!$F$9:$F${bkct_last},\'02.BKCT\'!$AC$9:$AC${bkct_last},$A{r},\'02.BKCT\'!$E$9:$E${bkct_last},$D$1))'
        ws.cell(r, 4).value = f'=IF($A{r}="",SUMIFS(\'03.BKPB\'!$F$6:$F${bkpb_last},\'03.BKPB\'!$B$6:$B${bkpb_last},$B{r},\'03.BKPB\'!$J$6:$J${bkpb_last},$D$1),SUMIFS(\'03.BKPB\'!$F$6:$F${bkpb_last},\'03.BKPB\'!$A$6:$A${bkpb_last},$A{r},\'03.BKPB\'!$J$6:$J${bkpb_last},$D$1))'
        ws.cell(r, 5).value = f"=C{r}-D{r}"
        ws.cell(r, 6).value = f'=IF($A{r}="",SUMIFS($E$4:$E${last},$B$4:$B${last},B{r}),SUMIFS($E$4:$E${last},$A$4:$A${last},A{r}))'
        ws.cell(r, 7).value = f'=IF(F{r}=0,"-","Check")'
        ws.cell(r, 8).value = f'=IF(A{r}="","Không tìm thấy Mã","")'
        for c in range(1, 9):
            ws.cell(r, c).alignment = Alignment(vertical="center")
        for c in range(3, 7):
            ws.cell(r, c).number_format = "#,##0"
        ws.cell(r, 6).font = Font(name="Calibri", size=11, bold=True, color="000000")

    ws["C2"].number_format = ws["D2"].number_format = ws["E2"].number_format = "#,##0"
    ws.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()
    red = Font(color="C00000")
    ws.conditional_formatting.add(f"A4:H{last}", FormulaRule(formula=["$A4=\"\""], font=red))

months: set[tuple[int, int]] = set()
for r in range(9, bkct_last + 1):
    text = str(bkct.cell(r, 2).value or "")
    match = re.search(r"/(\d{2})\.(\d{2})", text)
    if match:
        months.add((2000 + int(match.group(2)), int(match.group(1))))
period = f"{len(months)} tháng"
if len(months) == 1:
    year, month = next(iter(months))
    period = f"tháng {month:02d}/{year}"
for name in ("CheckChiTiet_2421", "CheckChiTiet_2422"):
    wb[name]["A1"] = f"BẢNG TỔNG HỢP ĐỐI CHIẾU CHI TIẾT - {period}"

wb["CheckChiTiet_2421"].sheet_properties.tabColor = "70AD47"
wb["CheckChiTiet_2422"].sheet_properties.tabColor = "FFC000"
for name in ("CheckChiTiet_2421", "CheckChiTiet_2422"):
    ws = wb[name]
    ws.auto_filter.ref = None
    for dim in ws.row_dimensions.values():
        dim.hidden = False
wb._sheets = [wb[name] for name in ("Inf", "CheckChiTiet_2421", "CheckChiTiet_2422", "02.BKCT", "03.BKPB")]

now = datetime.now()
info = wb["Inf"]
info["A9"] = f"Ngày giờ xuất file: {now:%d/%m/%Y %H:%M}"
info["A9"].alignment = Alignment(vertical="center")
filename = f"Check CCDC_{now:%Y%m%d_%H%M}.xlsx"
output = ROOT / filename
wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
wb.save(output)
(ROOT / "TOOL_Check_CCDC" / "_last_output.txt").write_text(filename, encoding="utf-8")
print(f"Đã tạo: {filename}")
