import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sqlite3, shutil, re
from datetime import datetime
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENTITY = "AS"
DATA_DIR = BASE_DIR

NUM = '#,##0'
THIN = Side(style='thin')
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BLUE = "2F75B5"
WHITE = "FFFFFF"
PAIR_COLORS = ["D6E4F0", "FCE4D6", "E2EFDA", "E8D5F5", "FFF2CC"]

VIET = str.maketrans({
    'à':'a','á':'a','ã':'a','ạ':'a','ả':'a',
    'ă':'a','ắ':'a','ằ':'a','ẳ':'a','ẵ':'a','ặ':'a',
    'â':'a','ấ':'a','ầ':'a','ẩ':'a','ẫ':'a','ậ':'a',
    'è':'e','é':'e','ẹ':'e','ẻ':'e','ẽ':'e',
    'ê':'e','ề':'e','ế':'e','ể':'e','ễ':'e','ệ':'e',
    'đ':'d',
    'ì':'i','í':'i','ị':'i','ỉ':'i','ĩ':'i',
    'ò':'o','ó':'o','ọ':'o','ỏ':'o','õ':'o',
    'ô':'o','ố':'o','ồ':'o','ổ':'o','ỗ':'o','ộ':'o',
    'ơ':'o','ớ':'o','ờ':'o','ở':'o','ỡ':'o','ợ':'o',
    'ù':'u','ú':'u','ụ':'u','ủ':'u','ũ':'u',
    'ư':'u','ứ':'u','ừ':'u','ử':'u','ữ':'u','ự':'u',
    'ỳ':'y','ý':'y','ỵ':'y','ỷ':'y','ỹ':'y',
})

def _find(folder, keywords):
    if not os.path.isdir(folder): return None
    for name in os.listdir(folder):
        plain = name.lower().translate(VIET)
        if all(k in plain for k in keywords):
            return os.path.join(folder, name)
    return None

def _f(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def _fmt(n):
    return round(float(n or 0), 0)

STRIP_PREFIXES = ["phân bổ chi phí công cụ dụng cụ:", "hạch toán chi phí chờ phân bổ:"]

def _strip_prefix(dien_giai):
    s = str(dien_giai or '').strip()
    for prefix in STRIP_PREFIXES:
        if s.lower().startswith(prefix):
            s = s[len(prefix):].strip()
            break
    return s

def _has_month_ref(s):
    return bool(re.search(r'(?:[Tt]\d{1,2}(?:\.\d{4})?|tháng\s+\d{1,2}(?:\s*[-–]\s*\d{1,2})?)', s))

def _strip_suffix(name):
    return re.sub(r'\s+[Tt]\d+\.\d{4}$', '', name).strip()

def _strip_suffix_display(name):
    s = str(name or '').strip()
    m = re.search(r'\s+[Tt](\d{1,2})(?:\.\d{4})?\s*$', s)
    if m:
        body = s[:m.start()]
        if _has_month_ref(body):
            return body.strip()
    return s

def _clean_dg_display(dien_giai):
    s = _strip_prefix(dien_giai)
    s = _strip_suffix_display(s)
    return s

def _clean_dg(dien_giai):
    s = _strip_prefix(dien_giai)
    s = _strip_suffix(s)
    return s

def _fmt_date(v):
    if hasattr(v, 'strftime'):
        return v.strftime('%d/%m/%Y')
    return str(v or '').strip()

def _extract_ccdc_name(dien_giai):
    s = _strip_prefix(dien_giai)
    if s == str(dien_giai or '').strip():
        return None
    for sep in [f" t{i}." for i in range(1, 13)]:
        idx = s.lower().find(sep)
        if idx > 0:
            s = s[:idx]
    return s.strip() or None

def _match_ccdc_item(ccdc_items, bk_name, bk_amount=0):
    if not bk_name: return None
    bn = bk_name.lower().strip()
    exact = []; prefix = []
    for item in ccdc_items:
        ten = (item['ten'] or '').lower().strip()
        if not ten: continue
        if bn == ten or ten == bn:
            exact.append(item)
        elif bn.startswith(ten) or ten.startswith(bn):
            prefix.append(item)
        else:
            bn_words = set(bn.replace('-',' ').split())
            ten_words = set(ten.replace('-',' ').split())
            common = bn_words & ten_words
            if len(common) >= 2 and len(common) >= min(len(bn_words), len(ten_words)) * 0.6:
                prefix.append(item)
    candidates = exact or prefix
    if not candidates: return None
    if bk_amount:
        for item in candidates:
            if abs((item['phan_bo'] or 0) - bk_amount) < 1000:
                return item
        return min(candidates, key=lambda x: abs((x['phan_bo'] or 0) - bk_amount))
    return candidates[0]


def _find_monthly(folder, keywords, month_code):
    if not os.path.isdir(folder): return None
    mc = month_code.lower()
    for name in os.listdir(folder):
        plain = name.lower().translate(VIET)
        if all(k in plain for k in keywords) and mc in plain:
            return os.path.join(folder, name)
    return None


def _process_month(month_code, ws, bcdps_data, TMP):
    conn = sqlite3.connect(os.path.join(TMP, f"ccdc_{month_code}.db"))
    mc = month_code.lower()
    f_ccdc = _find_monthly(BASE_DIR, ["tinh gia tri phan bo", "cong cu"], mc)
    f_cp   = _find_monthly(BASE_DIR, ["tong hop", "chi phi", "cho phan bo"], mc)
    f_2421 = _find_monthly(BASE_DIR, ["so chi tiet", "doi ung", "2421"], mc)
    f_2422 = _find_monthly(BASE_DIR, ["so chi tiet", "doi ung", "2422"], mc)
    if not all([f_ccdc, f_cp, f_2421, f_2422]):
        print(f"  [{month_code}] Thieu file - bo qua"); conn.close(); return
    def _scp(src):
        dst = os.path.join(TMP, os.path.basename(src))
        try: shutil.copy2(src, dst); return dst
        except: return src
    f_ccdc_cp=_scp(f_ccdc); f_cp_cp=_scp(f_cp); f_2421_cp=_scp(f_2421); f_2422_cp=_scp(f_2422)
    def _lw(p):
        wb = openpyxl.load_workbook(p, data_only=True); return wb, wb.active
    conn.execute("CREATE TABLE bcdps (tk TEXT, ten TEXT, du_no_dk REAL, du_co_dk REAL, ps_no REAL, ps_co REAL, du_no_ck REAL, du_co_ck REAL)")
    for tk_str, data in bcdps_data.items():
        conn.execute("INSERT INTO bcdps VALUES (?,?,?,?,?,?,?,?)",
                     (tk_str, data['ten'], data['du_no_dk'], data['du_co_dk'],
                      data['ps_no'], data['ps_co'], data['du_no_ck'], data['du_co_ck']))
    wb, ws_ccdc = _lw(f_ccdc_cp)
    conn.execute("CREATE TABLE ccdc (ma TEXT, ten TEXT, con_lai_dk REAL, phan_bo REAL, con_lai_ck REAL, tk_no TEXT, tk_co TEXT, tk_account TEXT, note TEXT)")
    for r in range(7, ws_ccdc.max_row + 1):
        if ws_ccdc.cell(r, 1).value is None: continue
        ma = str(ws_ccdc.cell(r, 1).value or '').strip()
        tk_no = str(ws_ccdc.cell(r, 18).value or '').strip()
        tk_co = str(ws_ccdc.cell(r, 19).value or '').strip()
        if tk_no in ('2421','2422'): acct = tk_no
        elif tk_co in ('2421','2422'): acct = tk_co
        elif tk_no == '242': acct = tk_no
        elif tk_co == '242': acct = tk_co
        else: acct = ''
        note = ''
        if acct == '' and _f(ws_ccdc.cell(r, 14).value) != 0:
            note = 'Check - TK N=C, can xac nhan'
        conn.execute("INSERT INTO ccdc VALUES (?,?,?,?,?,?,?,?,?)",
                     (ma, str(ws_ccdc.cell(r, 2).value or ''),
                      _f(ws_ccdc.cell(r, 10).value), _f(ws_ccdc.cell(r, 11).value),
                      _f(ws_ccdc.cell(r, 14).value), tk_no, tk_co, acct, note))
    wb.close()
    wb, ws_cp = _lw(f_cp_cp)
    conn.execute("CREATE TABLE cp (ma TEXT, ten TEXT, con_lai REAL, pb_trong_ky REAL, tk_no TEXT, tk_co TEXT)")
    for r in range(6, ws_cp.max_row + 1):
        if ws_cp.cell(r, 1).value is None: continue
        ma = str(ws_cp.cell(r, 1).value or '').strip()
        tk_co = str(ws_cp.cell(r, 11).value or '').strip()
        if tk_co not in ('242','2421','2422'): continue
        conn.execute("INSERT INTO cp VALUES (?,?,?,?,?,?)",
                     (ma, str(ws_cp.cell(r, 2).value or ''),
                      _f(ws_cp.cell(r, 9).value), _f(ws_cp.cell(r, 7).value),
                      str(ws_cp.cell(r, 10).value or '').strip(), tk_co))
    wb.close()
    conn.execute("CREATE TABLE so_242x (tk TEXT, ngay TEXT, so_ct TEXT, dien_giai TEXT, tk_doi_ung TEXT, ps_no REAL, ps_co REAL)")
    def _ls(p, tl):
        wb, ws_s = _lw(p)
        for r in range(9, ws_s.max_row + 1):
            sc = str(ws_s.cell(r, 2).value or '').strip()
            if not sc: continue
            conn.execute("INSERT INTO so_242x VALUES (?,?,?,?,?,?,?)",
                         (tl, _fmt_date(ws_s.cell(r, 1).value), sc, str(ws_s.cell(r, 3).value or ''),
                          str(ws_s.cell(r, 5).value or '').strip(),
                          _f(ws_s.cell(r, 6).value), _f(ws_s.cell(r, 7).value)))
        wb.close()
    _ls(f_2421_cp, "2421")
    _ls(f_2422_cp, "2422")
    conn.commit()

    b2421 = conn.execute("SELECT du_no_dk, du_co_dk, du_no_ck, ps_no, ps_co, ten FROM bcdps WHERE tk='2421'").fetchone() or (0,0,0,0,0,'')
    b2422 = conn.execute("SELECT du_no_dk, du_co_dk, du_no_ck, ps_no, ps_co, ten FROM bcdps WHERE tk='2422'").fetchone() or (0,0,0,0,0,'')
    ccdc_sum = conn.execute("SELECT tk_account, SUM(con_lai_ck), SUM(con_lai_dk), SUM(phan_bo) FROM ccdc GROUP BY tk_account").fetchall()
    cp_sum = conn.execute("SELECT tk_co, SUM(con_lai), SUM(pb_trong_ky) FROM cp GROUP BY tk_co").fetchall()
    ccdc_map = {r[0]: {"cl_ck":r[1], "cl_dk":r[2], "pb":r[3]} for r in ccdc_sum}
    cp_map = {r[0]: {"cl":r[1], "pb":r[2]} for r in cp_sum}

    cp_items_by_tk = {'2421': [], '2422': []}
    for r in conn.execute("SELECT ma, ten, pb_trong_ky, con_lai, tk_no, tk_co FROM cp WHERE tk_co IN ('2421','2422') AND ABS(pb_trong_ky) >= 1000 ORDER BY tk_co, ten"):
        ma, ten, pb, cl, tk_no, tk_co = r
        cp_items_by_tk[tk_co].append({'ma': ma, 'ten': str(ten or '').strip(), 'pb': _fmt(pb), 'cl': _fmt(cl), 'tk_no': tk_no})

    ccdc_items = [{"ma":r[0], "ten":r[1], "phan_bo":_f(r[2]), "tk_account":r[3]}
                  for r in conn.execute("SELECT ma, ten, phan_bo, tk_account FROM ccdc").fetchall()]
    ccdc_items_by_tk = {'2421': [], '2422': []}
    for r in conn.execute("SELECT ma, ten, phan_bo, con_lai_ck, tk_account, tk_no, tk_co FROM ccdc WHERE tk_account IN ('2421','2422') AND ABS(phan_bo) >= 1000 ORDER BY tk_account, ten"):
        ma, ten, pb, cl_ck, acct, tk_no, tk_co = r
        tk_du = tk_co if tk_no == acct else tk_no
        ccdc_items_by_tk[acct].append({'ma': ma, 'ten': str(ten or '').strip(), 'phan_bo': _fmt(pb), 'con_lai_ck': _fmt(cl_ck), 'tk_du': str(tk_du or '').strip()})

    all_ledger = list(conn.execute("SELECT tk, ngay, so_ct, dien_giai, tk_doi_ung, ps_no, ps_co FROM so_242x ORDER BY tk, ngay, so_ct"))
    entries_by_tk = {'2421': [], '2422': []}
    for r in all_ledger:
        tk, ngay, sc, dg, tdu, pn, pc = r
        pn=_fmt(pn); pc=_fmt(pc); cont=_fmt(pn-pc)
        if abs(cont)<1000: continue
        is_cp='PBCP' in sc.upper(); is_int=tdu in ('242','2421','2422')
        if is_cp: note='Phan bo CP'; ba=0; ca=0; cpa=abs(cont)
        elif is_int: note='Noi bo 242'; ba=0; ca=0; cpa=0
        else:
            note=''; ba=abs(cont); ca=0; cpa=0
            if 'PBCCDC' in sc.upper():
                nm=_extract_ccdc_name(dg)
                if nm:
                    mtch=_match_ccdc_item(ccdc_items, nm, ba)
                    if mtch:
                        ca=_fmt(mtch['phan_bo']); note='CCDC'
                        if abs(ca-ba)>=1000: note+=f' (SL={ca:,.0f})'
        entries_by_tk[tk].append({'tk_du':tdu,'tk_no':tk if pn>0 else tdu,'tk_co':tdu if pn>0 else tk,'so_ct':sc,'ngay':ngay,'dg':dg,'bk_amt':ba,'ccdc_amt':ca,'cp_amt':cpa,'contrib':cont,'note':note})

    close_data={}
    for tk in ('2421','2422'):
        bv=_fmt(b2421[2] if tk=='2421' else b2422[2]); cv=_fmt(ccdc_map.get(tk,{}).get('cl_ck',0)); pv=_fmt(cp_map.get(tk,{}).get('cl',0))
        close_data[tk]={'bcdps':bv,'ccdc':cv,'cp':pv,'lech':_fmt(bv-cv-pv)}

    ccdc_unassigned=[dict(ma=r[0],ten=r[1],cl_dk=_fmt(r[2]),pb=_fmt(r[3]),cl_ck=_fmt(r[4]),note=r[5])
                     for r in conn.execute("SELECT ma, ten, con_lai_dk, phan_bo, con_lai_ck, note FROM ccdc WHERE tk_account='' AND con_lai_ck != 0 ORDER BY ten")]

    def _mc(s): return _clean_dg(s).lower().strip()
    def _gp(es):
        gs={}
        for e in es:
            k=(_mc(e['dg']),str(e.get('tk_du','')).strip())
            if k not in gs: gs[k]={'total':0,'socp':[],'pe':e}
            gs[k]['total']+=-e['contrib']; gs[k]['socp'].append(e['so_ct'])
        return gs

    def _msn(items, af):
        mg={}
        for it in items:
            k=(_mc(it['ten']),str(it.get('tk_no') or it.get('tk_du') or '').strip())
            if k not in mg:
                mg[k]=dict(it); mg[k]['ma']=str(it.get('ma',''))
            else:
                cur=mg[k]; cur['ma']=','.join(x for x in [cur.get('ma',''),str(it.get('ma',''))] if x)
                for fld in af: cur[fld]=_fmt(cur.get(fld,0)+it.get(fld,0))
        return list(mg.values())

    am={}
    for tk in ('2421','2422'):
        ci=_msn(cp_items_by_tk[tk],('pb','cl')); lde=entries_by_tk[tk]
        pe=[e for e in lde if 'PBCP' in e['so_ct'].upper()]; pg=_gp(pe)
        ms=[]; upk=set()
        for cp_item in ci:
            ck=(_mc(cp_item['ten']),str(cp_item.get('tk_no','')).strip())
            if ck in pg and ck not in upk:
                upk.add(ck); g=pg[ck]; pt=g['total']; d=cp_item['pb']-pt if cp_item['pb']!=0 else -(cp_item['cl']-pt)
                ms.append({'src':'CP','ten':_clean_dg_display(cp_item['ten']),'cp_amt':cp_item['pb'],'pbcp_amt':pt,'diff':d,'note':'' if abs(d)<1000 else 'Lech tien','so_ct_ledger':','.join(g['socp']),'so_ct_file':cp_item['ma'],'tk':tk})
            else:
                d=cp_item['pb'] if cp_item['pb']!=0 else -cp_item['cl']
                ms.append({'src':'CP','ten':_clean_dg_display(cp_item['ten']),'cp_amt':cp_item['pb'],'pbcp_amt':0,'diff':d,'note':'Thieu PBCP','so_ct_ledger':'','so_ct_file':cp_item['ma'],'tk':tk})
        for k,g in pg.items():
            if k in upk: continue
            ms.append({'src':'PBCP','ten':_clean_dg_display(g['pe']['dg']),'cp_amt':0,'pbcp_amt':g['total'],'diff':-g['total'],'note':'Thua PBCP','so_ct_ledger':','.join(g['socp']),'so_ct_file':'','tk':tk})
        ccit=_msn(ccdc_items_by_tk.get(tk,[]),('phan_bo','con_lai_ck'))
        pbce=[e for e in lde if 'PBCCDC' in e['so_ct'].upper()]
        if pbce:
            pcg=_gp(pbce); upck=set()
            for cc in ccit:
                cck=(_mc(cc['ten']),str(cc.get('tk_du','')).strip())
                if cck in pcg and cck not in upck:
                    upck.add(cck); g=pcg[cck]; pt=g['total']; d=cc['phan_bo']-pt
                    ms.append({'src':'CCDC','ten':_clean_dg_display(cc['ten']),'cp_amt':cc['phan_bo'],'pbcp_amt':pt,'diff':d,'note':'' if abs(d)<1000 else 'Lech tien','so_ct_ledger':','.join(g['socp']),'so_ct_file':cc['ma'],'tk':tk})
                else:
                    ms.append({'src':'CCDC','ten':_clean_dg_display(cc['ten']),'cp_amt':cc['phan_bo'],'pbcp_amt':0,'diff':cc['phan_bo'],'note':'Con lai cuoi ky' if cc['phan_bo']==0 else 'Thieu PBCCDC','so_ct_ledger':'','so_ct_file':cc['ma'],'tk':tk})
            for k,g in pcg.items():
                if k in upck: continue
                ms.append({'src':'PBCCDC','ten':_clean_dg_display(g['pe']['dg']),'cp_amt':0,'pbcp_amt':g['total'],'diff':-g['total'],'note':'Thua PBCCDC','so_ct_ledger':','.join(g['socp']),'so_ct_file':'','tk':tk})
        else:
            for cc in ccit:
                ms.append({'src':'CCDC','ten':_clean_dg_display(cc['ten']),'cp_amt':cc['phan_bo'],'pbcp_amt':0,'diff':cc['phan_bo'],'note':'Con lai cuoi ky' if cc['phan_bo']==0 else 'Thieu PBCCDC','so_ct_ledger':'','so_ct_file':cc['ma'],'tk':tk})
        oth=[e for e in lde if 'PBCP' not in e['so_ct'].upper() and 'PBCCDC' not in e['so_ct'].upper()]
        ins=0; inc=0
        for oe in oth:
            if oe.get('note','')=='Noi bo 242': ins+=oe['contrib']; inc+=1
            else:
                kn=oe.get('note','') or f"TK {oe['tk_co'] if oe['tk_co'] else oe['tk_no']}"
                ms.append({'src':'Khac','ten':_clean_dg_display(oe['dg']),'cp_amt':0,'pbcp_amt':0,'diff':oe['contrib'],'note':kn,'so_ct_ledger':oe['so_ct'],'so_ct_file':'','tk':tk})
        if inc>0:
            ms.append({'src':'Noi bo','ten':f'Chuyen giua 2421 va 2422 ({inc} phieu)','cp_amt':0,'pbcp_amt':0,'diff':_fmt(ins),'note':'Noi bo 242 (khong can sua)','so_ct_ledger':'','so_ct_file':'','_netoff':'Noi bo 242','tk':tk})
        am[tk]=ms

    bc={}
    for tk in ('2421','2422'):
        for m in am[tk]:
            c=m.get('so_ct_file','')
            if c: bc.setdefault(c,[]).append(m)
    for code,items in bc.items():
        if len(items)>=2 and abs(sum(m['diff'] for m in items))<1000:
            for i,m in enumerate(items):
                m['_netoff']="Net off"
                m['_netoff_ct']=[(items[j]['tk'],items[j]['diff'],items[j]['ten']) for j in range(len(items)) if j!=i]
    for tk in ('2421','2422'):
        ml=am[tk]; act=[m for m in ml if not m.get('_netoff')]; bt={}
        for i,m in enumerate(act): bt.setdefault(m['ten'],[]).append(i)
        for ten,idxs in bt.items():
            pos=sorted([i for i in idxs if act[i]['diff']>0],key=lambda i: act[i]['diff'])
            neg=sorted([i for i in idxs if act[i]['diff']<0],key=lambda i: -act[i]['diff'])
            i=j=0
            while i<len(pos) and j<len(neg):
                pd=act[pos[i]]['diff']; nd=act[neg[j]]['diff']
                if abs(pd+nd)<1000:
                    act[pos[i]]['_netoff']="Net off cung TK"; act[neg[j]]['_netoff']="Net off cung TK"; i+=1; j+=1
                elif pd>-nd: j+=1
                else: i+=1
    ni=[]
    for tk in ('2421','2422'):
        for m in am[tk]:
            if m['src']=='Noi bo': ni.append(m)
    if len(ni)>=2 and abs(sum(m['diff'] for m in ni))<1000:
        for i,m in enumerate(ni):
            m['_netoff']="Noi bo 242"
            m['_netoff_ct']=[(ni[j]['tk'],ni[j]['diff'],ni[j]['ten']) for j in range(len(ni)) if j!=i]

    MC=10
    def _sec(row,text,color="D6E4F0"):
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=MC)
        c=ws.cell(row,1,text); c.fill=PatternFill("solid",fgColor=color)
        c.font=Font(name="Arial",bold=True,size=10); c.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[row].height=22
    def _hdr(row,values):
        for ci,v in enumerate(values,1):
            c=ws.cell(row,ci,v); c.fill=PatternFill("solid",fgColor=BLUE)
            c.font=Font(name="Arial",bold=True,size=9,color=WHITE); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=GRID
    def _rw(row,values,fmt_cols=None,bold=False,fill=None):
        for ci,v in enumerate(values,1):
            c=ws.cell(row,ci,v); c.font=Font(name="Arial",size=9,bold=bold)
            c.alignment=Alignment(horizontal="right" if (fmt_cols and ci in fmt_cols) else "left",vertical="center",wrap_text=True); c.border=GRID
            if fmt_cols and ci in fmt_cols: c.number_format=NUM
            if fill: c.fill=PatternFill("solid",fgColor=fill)

    mlabels={'t1':'Thang 1','t2':'Thang 2','t3':'Thang 3','t4':'Thang 4','t5':'Thang 5','t6':'Thang 6'}
    ml=mlabels.get(month_code.lower(),month_code.upper())
    ws.merge_cells("A1:J1")
    ws["A1"]=f"PHIEU CAN DIEU CHINH - {ml} - TK 242 ({ENTITY})"; ws["A1"].font=Font(name="Arial",size=13,bold=True,color=BLUE)
    ws.merge_cells("A2:J2"); ws["A2"]=f"Xuat luc: {datetime.now():%d/%m/%Y %H:%M}"; ws["A2"].font=Font(name="Arial",size=9,color="7F7F7F")
    row=4
    for tk in ('2421','2422'):
        _sec(row,f"TK {tk}: BCDPS.cl={close_data[tk]['bcdps']:>,.0f} - CCDC.cl={close_data[tk]['ccdc']:>,.0f} - CP.cl={close_data[tk]['cp']:>,.0f} = Lech {close_data[tk]['lech']:>,.0f}")
        row+=1
    if ccdc_unassigned:
        row+=1; _sec(row,"CCDC CHUA PHAN BO","FCE4D6"); row+=1
        _hdr(row,["STT","Nguon","Dien giai","BCĐPS","CCDC","CP","Tong","Lech","Ket luan"]); row+=1
        for u in ccdc_unassigned:
            _rw(row,[1,'',str(u['ten'])[:60],0,u['cl_ck'],0,u['cl_ck'],-u['cl_ck'],u['note']],fmt_cols={4,5,6,7,8},fill='FFF2CC'); row+=1
        row+=1
    row+=1; _sec(row,f"PHAN TICH {ml.upper()}","E2EFDA"); row+=1
    _hdr(row,["STT","Nguon","Dien giai","So CT CP","So CT PBCP","CP Amt","PBCP Amt","Lech","Can sua"])
    for ci,w in {1:5,2:8,3:55,4:16,5:16,6:16,7:16,8:16,9:20}.items(): ws.column_dimensions[chr(64+ci) if ci<=26 else ''].width=w
    row+=1
    mt=am['2421']+am['2422'] if '2421' in am and '2422' in am else []
    stt=0
    nf=sorted([m for m in mt if m.get('src','') not in ('','DK') and not m.get('_netoff') and abs(m['diff'])>=1000],key=lambda x:x.get('src',''))
    if nf:
        _sec(row,"CAN SUA","FCE4D6"); row+=1; cs=row
        for m in nf:
            stt+=1; ws.row_dimensions[row].height=28
            _rw(row,[stt,m['src'],str(m['ten'])[:80],m.get('so_ct_file',''),m.get('so_ct_ledger',''),m['cp_amt'],m['pbcp_amt'],m['diff'],m.get('note','')],fmt_cols={6,7,8},fill="FCE4D6")
            row+=1
        _rw(row,['','',"Tong can sua",'','',0,0,f"=SUM(H{cs}:H{row-1})",''],bold=True,fmt_cols={6,7,8}); row+=1
    noi=sorted([m for m in mt if m.get('_netoff')],key=lambda x:x.get('src',''))
    if noi:
        _sec(row,"NET OFF","FFF2CC"); row+=1; ns=row
        for m in noi:
            stt+=1; ws.row_dimensions[row].height=28
            _rw(row,[stt,m['src'],str(m['ten'])[:80],m.get('so_ct_file',''),m.get('so_ct_ledger',''),m['cp_amt'],m['pbcp_amt'],m['diff'],m.get('_netoff','')],fmt_cols={6,7,8})
            row+=1
            if m.get('_netoff_ct'):
                for ct_tk,ct_d,_ in m['_netoff_ct']:
                    _rw(row,['','',f"  -> TK {ct_tk}: {ct_d:+,}",'','',0,0,0,''],fmt_cols={6,7,8}); row+=1
        _rw(row,['','',"Tong net off",'','',0,0,f"=SUM(H{ns}:H{row-1})",''],bold=True,fmt_cols={6,7,8}); row+=1
    for tk in ('2421','2422'):
        print(f"  [{month_code}] {tk}: lech={close_data[tk]['lech']:>12,.0f} = BCDPS={close_data[tk]['bcdps']:>12,.0f} - CCDC={close_data[tk]['ccdc']:>12,.0f} - CP={close_data[tk]['cp']:>12,.0f}")
    conn.close()


def main():
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    TMP = os.path.join(BASE_DIR, "~tmp")
    if os.path.exists(TMP):
        shutil.rmtree(TMP, ignore_errors=True)
    os.makedirs(TMP, exist_ok=True)
    f_bcdps = _find(BASE_DIR, ["can doi", "phat sinh"])
    if not f_bcdps:
        print("[LOI] Khong tim thay BCDPS"); return
    wb = openpyxl.load_workbook(f_bcdps, data_only=True)
    ws = wb.active
    bcdps_data = {}
    for r in range(10, ws.max_row + 1):
        tk = ws.cell(r, 1).value
        if tk is not None:
            ts = str(tk).strip()
            bcdps_data[ts] = {'du_no_dk': _f(ws.cell(r,4).value), 'du_co_dk': _f(ws.cell(r,5).value),
                              'ps_no': _f(ws.cell(r,6).value), 'ps_co': _f(ws.cell(r,7).value),
                              'du_no_ck': _f(ws.cell(r,10).value), 'du_co_ck': _f(ws.cell(r,11).value),
                              'ten': str(ws.cell(r,2).value or '')}
    wb.close()
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    months = [("T1","Thang 1"),("T2","Thang 2"),("T3","Thang 3"),
              ("T4","Thang 4"),("T5","Thang 5"),("T6","Thang 6")]
    for mc, ml in months:
        print(f"\n  === Processing {ml} ===")
        ws = wb_out.create_sheet(ml)
        _process_month(mc, ws, bcdps_data, TMP)
    out = os.path.join(BASE_DIR, f"z_DoiChieu_CCDC_{ENTITY}_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    wb_out.save(out)
    wb_out.close()
    try: shutil.rmtree(TMP, ignore_errors=True)
    except: pass
    print(f"\n  Xuat xong: {out}")
    return out

if __name__ == "__main__":
    main()
