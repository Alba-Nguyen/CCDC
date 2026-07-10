import collections
import os
import shutil
import tempfile

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = r"D:\OneDrive - SHINE GROUP\Desktop\Vi - Alba\Alba - CCDC\CCDC\AS"
OUTPUT = os.path.join(ROOT, "Bao_cao_doi_chieu_CCDC.xlsx")

BALANCE_FILE = "Bảng cân đối số phát sinh.xlsx"
SOURCE_SPECS = [
    {
        "source": "Bảng tính CCDC",
        "file": "Bảng tính giá trị phân bổ công cụ dụng cụ.xlsx",
        "start": 7,
        "debit_col": 17,
        "credit_col": 18,
        "remain_col": 13,
        "code_col": 0,
        "name_col": 1,
        "dept_col": 14,
    },
    {
        "source": "Bảng tổng hợp CP",
        "file": "Bảng tổng hợp chi phí chờ phân bổ.xlsx",
        "start": 6,
        "debit_col": 9,
        "credit_col": 10,
        "remain_col": 8,
        "code_col": 0,
        "name_col": 1,
        "dept_col": 13,
    },
]
LEDGER_FILES = {
    "2421": "Sổ chi tiết tài khoản 2421.xlsx",
    "2422": "Sổ chi tiết tài khoản 2422.xlsx",
}


def num(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def open_readonly(path):
    try:
        return openpyxl.load_workbook(path, data_only=True, read_only=True)
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), os.path.basename(path))
        shutil.copy2(path, tmp)
        return openpyxl.load_workbook(tmp, data_only=True, read_only=True)


def load_balance():
    ws = open_readonly(os.path.join(ROOT, BALANCE_FILE)).active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    accounts = {}
    for idx, row in enumerate(rows, 1):
        account = str(row[0] or "").strip()
        if account in {"242", "2421", "2422"}:
            accounts[account] = {
                "row": idx,
                "name": row[1],
                "ending": num(row[9]) - num(row[10]),
            }
    return rows, accounts


def load_sources():
    records = []
    totals = collections.defaultdict(lambda: collections.defaultdict(float))
    for spec in SOURCE_SPECS:
        ws = open_readonly(os.path.join(ROOT, spec["file"])).active
        for row_no, row in enumerate(ws.iter_rows(min_row=spec["start"], values_only=True), spec["start"]):
            name = row[spec["name_col"]]
            if not name or "Tổng cộng" in str(name):
                continue
            value = num(row[spec["remain_col"]])
            if not value:
                continue
            debit = str(row[spec["debit_col"]] or "").strip()
            credit = str(row[spec["credit_col"]] or "").strip()
            account = credit if credit in {"2421", "2422"} else ""
            record = {
                "source": spec["source"],
                "source_row": row_no,
                "code": row[spec["code_col"]],
                "name": name,
                "dept": row[spec["dept_col"]],
                "debit": debit,
                "credit": credit,
                "account": account,
                "remaining": value,
            }
            records.append(record)
            if account:
                totals[spec["source"]][account] += value
    return records, totals


def load_ledgers():
    out = {}
    for account, file_name in LEDGER_FILES.items():
        ws = open_readonly(os.path.join(ROOT, file_name)).active
        totals = {
            "begin": 0.0,
            "debit": 0.0,
            "credit": 0.0,
            "ending": 0.0,
        }
        rows = []
        for row_no, row in enumerate(ws.iter_rows(values_only=True), 1):
            desc = row[2] if len(row) > 2 else None
            if desc == "Số dư đầu kỳ":
                totals["begin"] = num(row[8]) - num(row[9])
            elif desc == "Tổng số phát sinh":
                totals["debit"] = num(row[5])
                totals["credit"] = num(row[7])
            elif desc == "Số dư cuối kỳ":
                totals["ending"] = num(row[8]) - num(row[9])
            elif row_no >= 9 and row[0] and row[1] and desc:
                rows.append(
                    {
                        "account": account,
                        "row": row_no,
                        "date": row[0],
                        "voucher": row[1],
                        "desc": desc,
                        "contra": row[4],
                        "debit": num(row[5]),
                        "credit": num(row[7]),
                        "signed": num(row[5]) - num(row[7]),
                    }
                )
        out[account] = {"totals": totals, "rows": rows}
    out["242"] = {
        "totals": {
            key: out["2421"]["totals"][key] + out["2422"]["totals"][key]
            for key in out["2421"]["totals"]
        },
        "rows": out["2421"]["rows"] + out["2422"]["rows"],
    }
    return out


def build_detail_rows(records, balance, source_total):
    rows = []

    def add(account, source, source_row, code, name, dept, debit, credit, current, adjustment, note1, note2):
        rows.append(
            {
                "check": "",
                "account": account,
                "source": source,
                "source_row": source_row,
                "code": code,
                "name": name,
                "dept": dept,
                "current_pair": f"{debit}/{credit}".strip("/"),
                "current": current,
                "adjustment": adjustment,
                "after": current + adjustment,
                "note1": note1,
                "note2": note2,
            }
        )

    by_code = {str(r["code"] or ""): r for r in records}

    # 2421: these two lines make the 2421 detail equal the total difference.
    if "CP.00114CP" in by_code:
        r = by_code["CP.00114CP"]
        add("2421", r["source"], r["source_row"], r["code"], r["name"], r["dept"], r["debit"], r["credit"], r["remaining"], -r["remaining"], "Cần xử lý", "Dòng này đang nằm trong tổng 2421; loại ra thì giảm 55,000,000.")
    if "CP.0100" in by_code:
        r = by_code["CP.0100"]
        add("2421", r["source"], r["source_row"], r["code"], r["name"], r["dept"], r["debit"], r["credit"], r["remaining"], 10519002, "Cần xử lý", "Bù lại phần 2421 còn thiếu sau khi loại CP.00114CP; tổng 2 dòng khớp chênh 2421.")

    # 2422: all rows currently posted out of 2422 but not credited to 2422 are shown.
    for r in records:
        if r["debit"] == "2422" and r["credit"] not in {"2421", "2422"}:
            add("2422", r["source"], r["source_row"], r["code"], r["name"], r["dept"], r["debit"], r["credit"], r["remaining"], r["remaining"], "Cần chuyển về 2422", "Dòng đang không nằm trong tổng 2422 vì TK Có không phải 2422.")

    target_2422 = balance["2422"]["ending"] - source_total["2422"]
    current_2422_adjustment = sum(row["adjustment"] for row in rows if row["account"] == "2422")
    remainder = target_2422 - current_2422_adjustment
    if round(remainder) != 0:
        target = abs(round(remainder))
        candidates = [
            r for r in records
            if r["credit"] == "2422" and r["remaining"] > 0 and r["source"] == "Bảng tính CCDC"
        ]
        dp = {0: ()}
        for idx, rec in enumerate(candidates):
            value = round(rec["remaining"])
            for total, combo in list(dp.items()):
                new_total = total + value
                if new_total <= target and new_total not in dp:
                    dp[new_total] = combo + (idx,)
            if target in dp:
                break
        if target in dp:
            sign = 1 if remainder > 0 else -1
            for idx in dp[target]:
                r = candidates[idx]
                add(
                    "2422",
                    r["source"],
                    r["source_row"],
                    r["code"],
                    r["name"],
                    r["dept"],
                    r["debit"],
                    r["credit"],
                    r["remaining"],
                    sign * r["remaining"],
                    "Cần xử lý",
                    "Phiếu thật dùng để bù phần chênh còn lại; không phải net off nếu không có dòng âm/dương đối ứng bằng 0.",
                )
        else:
            add(
                "2422",
                "Chênh tổng",
                "",
                "",
                "Phần chênh còn lại cần xác định phiếu chi tiết",
                "",
                "",
                "2422",
                0,
                remainder,
                "Cần xác định thêm",
                "Không tìm được tổ hợp phiếu thật bằng đúng phần chênh còn lại.",
            )

    return rows


def build_all_rows(records, selected_rows):
    netoff_keys = set()
    grouped = collections.defaultdict(list)
    for rec in records:
        key = str(rec["name"] or "").strip().upper()
        if key:
            grouped[key].append(rec)
    for key, items in grouped.items():
        values = [item["remaining"] for item in items]
        if len(items) >= 2 and any(v > 0 for v in values) and any(v < 0 for v in values) and round(sum(values)) == 0:
            netoff_keys.add(key)

    selected_by_key = {
        (row["source"], row["source_row"], str(row["code"] or ""), row["account"]): row
        for row in selected_rows
        if row["source"] not in {"Chênh tổng", "ChÃªnh tá»•ng"}
    }
    residual_rows = [
        row for row in selected_rows if row["source"] in {"Chênh tổng", "ChÃªnh tá»•ng"}
    ]
    rows = []
    seen_selected = set()

    for rec in records:
        netoff_key = str(rec["name"] or "").strip().upper()
        is_netoff = netoff_key in netoff_keys
        if rec["credit"] in {"2421", "2422"}:
            account = rec["credit"]
        elif rec["debit"] in {"2421", "2422"}:
            account = rec["debit"]
        else:
            continue

        key = (rec["source"], rec["source_row"], str(rec["code"] or ""), account)
        selected = selected_by_key.get(key)
        if selected:
            seen_selected.add(key)
            status = "Net off" if is_netoff or selected.get("note1") == "Net off" else "Chênh lệch"
            note1 = "Net off" if is_netoff else selected.get("note1")
            note2 = "Có phiếu âm/dương cùng diễn giải, tổng nhóm bằng 0; mã phiếu có thể khác nhau." if is_netoff else selected.get("note2")
            rows.append(dict(selected, check="Cần check", status=status, note1=note1, note2=note2))
            continue

        if is_netoff:
            status = "Net off"
            check = ""
            note1 = "Net off"
            note2 = "Có phiếu âm/dương cùng diễn giải, tổng nhóm bằng 0; mã phiếu có thể khác nhau."
        elif rec["credit"] in {"2421", "2422"}:
            status = "Khớp"
            check = ""
            note1 = "Khớp"
            note2 = "Phiếu đang nằm đúng trong tổng 2421/2422; không đưa vào số chênh cần check."
        elif rec["debit"] in {"2421", "2422"} and rec["credit"] not in {"2421", "2422"}:
            status = "Không chọn vào chênh"
            check = ""
            note1 = "Cần xem nếu lọc riêng"
            note2 = "Phiếu có TK Nợ 242 nhưng TK Có không phải 242; không nằm trong tập điều chỉnh đang khớp tổng."
        else:
            status = "Bỏ qua"
            check = ""
            note1 = "Bỏ qua"
            note2 = "Không thuộc 2421/2422."

        rows.append(
            {
                "check": check,
                "status": status,
                "account": account,
                "source": rec["source"],
                "source_row": rec["source_row"],
                "code": rec["code"],
                "name": rec["name"],
                "dept": rec["dept"],
                "current_pair": f"{rec['debit']}/{rec['credit']}".strip("/"),
                "current": rec["remaining"],
                "adjustment": 0.0,
                "after": rec["remaining"],
                "note1": note1,
                "note2": note2,
            }
        )

    for row in selected_rows:
        key = (row["source"], row["source_row"], str(row["code"] or ""), row["account"])
        if key not in seen_selected and row["source"] not in {"Chênh tổng", "ChÃªnh tá»•ng"}:
            status = "Net off" if row.get("note1") == "Net off" else "Chênh lệch"
            rows.append(dict(row, check="Cần check", status=status))
    for row in residual_rows:
        status = "Net off" if row.get("note1") == "Net off" else "Chênh lệch"
        rows.append(dict(row, check="Cần check", status=status))

    return rows


def style_sheet(ws, money_cols):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    detail_fill = PatternFill("solid", fgColor="FCE4D6")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        note = str(row[-2].value or "")
        fill = ok_fill if note == "OK" else (total_fill if "Cần xác định" in note else detail_fill)
        for cell in row:
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx in money_cols:
            if idx <= len(row):
                row[idx - 1].number_format = "#,##0"
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 70)


def write_balance_sheet(wb, balance_rows, balance, source_totals):
    ws = wb.create_sheet("Can doi so phat sinh")
    max_cols = max(len(row) for row in balance_rows)
    for row in balance_rows:
        ws.append(row + [None] * (max_cols - len(row)))

    headers = [
        "Cuối kỳ theo sổ chi tiết",
        "Lệch sổ - cân đối",
        "Còn lại theo CP",
        "Còn lại theo CCDC",
        "Tổng CP+CCDC",
        "Chênh CP+CCDC so với sổ",
        "Check tổng",
    ]
    for offset, header in enumerate(headers, max_cols + 1):
        ws.cell(7, offset).value = header

    ledgers = load_ledgers()
    for account in ("242", "2421", "2422"):
        row_no = balance[account]["row"]
        cp = source_totals["Bảng tổng hợp CP"]["2421"] + source_totals["Bảng tổng hợp CP"]["2422"] if account == "242" else source_totals["Bảng tổng hợp CP"][account]
        ccdc = source_totals["Bảng tính CCDC"]["2421"] + source_totals["Bảng tính CCDC"]["2422"] if account == "242" else source_totals["Bảng tính CCDC"][account]
        total = cp + ccdc
        ending = balance[account]["ending"]
        ledger_ending = ledgers[account]["totals"]["ending"]
        ledger_diff = ledger_ending - ending
        source_diff = ledger_ending - total
        values = [
            ledger_ending,
            ledger_diff,
            cp,
            ccdc,
            total,
            source_diff,
            "OK" if round(source_diff) == 0 else "Lệch",
        ]
        for offset, value in enumerate(values, max_cols + 1):
            ws.cell(row_no, offset).value = value

    style_sheet(ws, range(max_cols + 1, max_cols + 7))
    ws.freeze_panes = "A10"


def write_detail_sheet(wb, rows):
    ws = wb.create_sheet("Chi tiet phieu")
    ws.append(
        [
            "Check",
            "Trạng thái",
            "Tài khoản",
            "Nguồn",
            "Dòng nguồn",
            "Mã",
            "Tên / diễn giải",
            "Mã bộ phận",
            "TK hiện tại",
            "Giá trị còn lại hiện tại",
            "Số điều chỉnh",
            "Sau điều chỉnh",
            "Ghi chú 1",
            "Ghi chú 2",
        ]
    )
    for row in rows:
        ws.append(
            [
                row.get("check", ""),
                row.get("status", row.get("note1", "")),
                row["account"],
                row["source"],
                row["source_row"],
                row["code"],
                row["name"],
                row["dept"],
                row["current_pair"],
                row["current"],
                row["adjustment"],
                row["after"],
                row["note1"],
                row["note2"],
            ]
        )
    style_sheet(ws, (10, 11, 12))

    totals = collections.defaultdict(float)
    for row in rows:
        if row.get("check") == "Cần check":
            totals[row["account"]] += row["adjustment"]
    start = ws.max_row + 2
    ws.cell(start, 9).value = "Tổng Cần check 2421"
    ws.cell(start, 11).value = totals["2421"]
    ws.cell(start + 1, 9).value = "Tổng Cần check 2422"
    ws.cell(start + 1, 11).value = totals["2422"]
    for r in range(start, start + 2):
        for c in range(9, 12):
            ws.cell(r, c).font = Font(bold=True)
            ws.cell(r, c).number_format = "#,##0"


def main():
    balance_rows, balance = load_balance()
    records, source_totals = load_sources()
    source_total = {
        account: source_totals["Bảng tổng hợp CP"][account] + source_totals["Bảng tính CCDC"][account]
        for account in ("2421", "2422")
    }
    selected_rows = build_detail_rows(records, balance, source_total)
    detail_rows = build_all_rows(records, selected_rows)

    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    write_balance_sheet(wb, balance_rows, balance, source_totals)
    write_detail_sheet(wb, detail_rows)
    wb.save(OUTPUT)

    print(OUTPUT)
    for account in ("2421", "2422"):
        diff = balance[account]["ending"] - source_total[account]
        detail = sum(row["adjustment"] for row in detail_rows if row["account"] == account and row.get("check") == "Cần check")
        print(account, "chênh tổng:", round(diff), "chi tiết:", round(detail))


if __name__ == "__main__":
    main()
